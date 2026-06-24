from __future__ import annotations

import json
import os
import re
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def load_env(path: str) -> None:
    for line in Path(path).read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key, value.strip().strip('"').strip("'"))


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def display(value: object) -> str:
    return str(value or "").strip()


def split_emails(value: object) -> set[str]:
    return {clean(part) for part in re.split(r"[;,]", str(value or "")) if clean(part)}


def normalize_class(value: object) -> str:
    label = display(value)
    match = re.match(r"^(\d)-(\d+)$", label)
    if match:
        return f"0{match.group(1)}-{match.group(2)}"
    return label


def supabase_get(path: str) -> list[dict]:
    base_url = os.environ["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
    key = os.environ["SUPABASE_SECRET_KEY"]
    request = urllib.request.Request(f"{base_url}{path}")
    request.add_header("apikey", key)
    request.add_header("Authorization", f"Bearer {key}")
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def main() -> None:
    load_env(".env.local")
    workbook = load_workbook(
        "input/Alle_Schüler_2025-26.xlsx",
        read_only=True,
        data_only=True,
    )
    sheet = workbook["Schueler"]

    current_rows = []
    current_email_tokens = set()
    for first_name, last_name, email, school_year, school_class in sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        if not any([first_name, last_name, email, school_class]):
            continue
        email_tokens = split_emails(email)
        current_rows.append(
            {
                "first_name": display(first_name),
                "last_name": display(last_name),
                "emails": sorted(email_tokens),
                "raw_email": display(email),
                "class": normalize_class(school_class),
                "name_key": (clean(first_name), clean(last_name)),
            }
        )
        current_email_tokens.update(email_tokens)

    people = supabase_get(
        "/rest/v1/person"
        "?select=id,legacy_user_id,first_name,last_name,email,jahrgang"
        "&person_type=eq.schueler&status=eq.aktiv"
        "&order=last_name.asc&limit=1000"
    )
    assignments = supabase_get(
        "/rest/v1/person_class_assignment"
        "?select=person_id,school_class:school_class_id(label)"
        "&valid_until=is.null&limit=1000"
    )
    class_by_person = {}
    for assignment in assignments:
        school_class = assignment.get("school_class")
        if isinstance(school_class, list):
            school_class = school_class[0] if school_class else None
        class_by_person[assignment["person_id"]] = (
            school_class.get("label") if school_class else None
        )

    current_name_counts = Counter(row["name_key"] for row in current_rows)
    duplicate_names = [
        {
            "name": f"{key[1]}, {key[0]}",
            "count": count,
        }
        for key, count in sorted(current_name_counts.items())
        if count > 1
    ]

    extra_by_email = []
    for person in people:
        email = clean(person.get("email"))
        if email in current_email_tokens:
            continue
        extra_by_email.append(
            {
                "legacy_user_id": person.get("legacy_user_id"),
                "name": f"{display(person.get('last_name'))}, {display(person.get('first_name'))}",
                "email": display(person.get("email")),
                "class": class_by_person.get(person["id"]) or "-",
                "jahrgang": person.get("jahrgang"),
            }
        )

    print(
        json.dumps(
            {
                "current_file_rows": len(current_rows),
                "current_file_email_tokens": len(current_email_tokens),
                "supabase_active_students": len(people),
                "extra_by_email_count": len(extra_by_email),
                "extra_by_email": extra_by_email,
                "duplicate_names_in_current_file": duplicate_names,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
